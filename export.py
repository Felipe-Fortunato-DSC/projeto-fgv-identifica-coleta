"""
Exportação para Excel:
  - export_to_excel: .xlsx (Open XML, openpyxl) para produtos selecionados.
  - export_carga_real: .xls (Excel 97-2003, BIFF8, xlwt) para Carga BP,
    formato exigido pelo sistema legado.
"""

import io
from datetime import date

import pandas as pd
import xlwt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Mapeamento coluna -> rótulo amigável (mesmo da exibição)
COLUMN_LABELS = {
    "data_coleta": "Data Coleta",
    "plataforma": "Plataforma",
    "cod_informante": "Cód. Informante",
    "nome_informante": "Nome Informante",
    "periodicidade": "Periodicidade",
    "tipo_preco": "Tipo Preço",
    "cod_insumo": "Cód. Insumo",
    "ean": "EAN",
    "sku": "SKU",
    "insumo_informado": "Insumo Informado",
    "url": "URL",
    "descricao": "Descrição",
    "marca": "Marca",
    "uf": "UF",
    "moeda": "Moeda",
    "preco": "Preço (R$)",
    "preco_promocional": "Preço Promo (R$)",
    "id_produto": "ID Produto",
    "id_coleta": "ID Coleta",
}

HEADER_BG = "1B4F72"   # azul escuro
HEADER_FG = "FFFFFF"   # branco


def _auto_width(ws) -> None:
    """Ajusta a largura de cada coluna ao conteúdo mais longo."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def export_to_excel(df: pd.DataFrame) -> bytes:
    """
    Recebe um DataFrame com os produtos selecionados e retorna
    o conteúdo do arquivo .xlsx como bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "produtos"

    header_font = Font(bold=True, color=HEADER_FG)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    center = Alignment(horizontal="center", vertical="center")

    # --- Cabeçalho ---
    headers = [COLUMN_LABELS.get(c, c) for c in df.columns]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    ws.row_dimensions[1].height = 20

    # --- Dados ---
    price_cols = {"preco", "preco_promocional"}
    date_cols = {"data_coleta"}

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, (col_name, value) in enumerate(
            zip(df.columns, row), start=1
        ):
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_name in price_cols and value is not None:
                try:
                    cell.value = float(value)
                    cell.number_format = 'R$ #,##0.00'
                except (ValueError, TypeError):
                    cell.value = value
            elif col_name in date_cols and isinstance(value, date):
                cell.value = value
                cell.number_format = "DD/MM/YYYY"
            else:
                cell.value = value

            cell.alignment = Alignment(vertical="center")

    _auto_width(ws)

    # Freeze primeira linha
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# Layout da Carga BP (real, espelhando coleta_*.xls).
# Cada item: (rótulo no Excel, origem no parquet OU None, tipo).
# tipo ∈ {"text", "price", "date"}; None em origem = coluna em branco.
CARGA_REAL_COLUMNS: list[tuple[str, str | None, str]] = [
    ("JOB",                                  "JOB",                      "text"),
    ("Insumo Informado",                     "NR_SEQ_INSINF",            "text"),
    ("Sinônimo Insumo Informado",            "DS_SINO_NOME_INS_INSINF",  "text"),
    ("Código do Informante",                 "CD_INFORM",                "text"),
    ("Código do Insumo",                     "CD_INSUMO",                "text"),
    ("Tipo de Preço",                        "CD_TPPRECO",               "text"),
    ("Pais",                                 "PAIS",                     "text"),
    ("Região",                               "REGIAO",                   "text"),
    ("Estado",                               "ESTADO",                   "text"),
    ("Municipio",                            "MUNICIPIO",                "text"),
    ("Bairro",                               "BAIRRO",                   "text"),
    ("Pais_Retirada",                        "PAIS",                     "text"),
    ("Regiao_Retirada",                      "REGIAO",                   "text"),
    ("Estado_Retirada",                      "ESTADO",                   "text"),
    ("Municipio_Retirada",                   "MUNICIPIO",                "text"),
    ("Bairro_Retirada",                      None,                       "text"),
    ("Cotação",                              "CD_COTACAO",               "text"),
    ("Periodicidade",                        "CD_PERIOD",                "text"),
    ("Data do Preço",                        "DT_COL_PRECCOL",           "date"),
    ("Data Prevista",                        "DATA_PREVISTA",            "date"),
    ("Valor do Preço",                       "VL_PRECCOL",               "price"),
    ("Moeda",                                "CD_MOEDA",                 "text"),
    ("Preço Promocional",                    None,                       "price"),
    ("Valor do Frete",                       None,                       "price"),
    ("Taxa do Frete",                        "TAXA_FRETE",               "price"),
    ("Frete Incluso",                        None,                       "text"),
    ("Frete Nao Declarado",                  None,                       "text"),
    ("Valor do Desconto",                    None,                       "price"),
    ("Taxa do Desconto",                     None,                       "price"),
    ("Desconto Incluso",                     None,                       "text"),
    ("Desconto Não Declarado",               None,                       "text"),
    ("Coletor Padrão",                       "CD_COLETOR",               "text"),
    ("FT",                                   None,                       "text"),
    ("Justificativa Livre",                  "DS_OBS_PRECCOL",           "text"),
    ("URL Insumo Informado",                 "URL_DO_INSUMO",            "text"),
    ("Arquivo com Preço",                    None,                       "text"),
    ("Nome Insumo",                          "NM_INSUMO",                "text"),
    ("Característica Insumo",                "DS_INSUMO",                "text"),
    ("Especificação Insumo",                 None,                       "text"),
    ("Marca Insumo",                         "CD_MARCFAB",               "text"),
    ("Embalagem Insumo",                     "CD_EMB",                   "text"),
    ("Quantidade Insumo",                    "QT_MED_INSUMO",            "text"),
    ("Unidade Medida Insumo",                "CD_MEDIDA",                "text"),
    ("Grupo de Coleta Insumo",               "GRUPO_DE_COLETA",          "text"),
    ("Obs Insumo",                           "DS_OBS_INSINF",            "text"),
    ("EAN Insumo",                           "EAN",                      "text"),
]


def _to_blank(value) -> bool:
    """True se o valor deve virar célula vazia (None, NaN, pd.NA, NaT, string vazia)."""
    if value is None:
        return True
    # pd.isna captura np.nan, pd.NA, pd.NaT — em scalar retorna bool.
    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _xls_header_style() -> xlwt.XFStyle:
    style = xlwt.XFStyle()
    font = xlwt.Font(); font.bold = True; font.colour_index = 0x01  # branco
    style.font = font
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern.pattern_fore_colour = 0x12  # azul escuro padrão da paleta BIFF
    style.pattern = pattern
    align = xlwt.Alignment()
    align.horz = xlwt.Alignment.HORZ_CENTER
    align.vert = xlwt.Alignment.VERT_CENTER
    style.alignment = align
    return style


def _xls_style(num_format: str | None = None) -> xlwt.XFStyle:
    style = xlwt.XFStyle()
    if num_format:
        style.num_format_str = num_format
    align = xlwt.Alignment()
    align.vert = xlwt.Alignment.VERT_CENTER
    style.alignment = align
    return style


def _write_carga_sheet(
    ws,
    df_enc: pd.DataFrame,
    layout: list[tuple[str, str | None, str]],
) -> None:
    """Escreve cabeçalho + linhas no worksheet xlwt seguindo o layout dado.
    Mantém formatação de preço/data/texto, larguras automáticas e header
    estilizado. Extraído para suportar a carga "reprovada" com colunas extras.
    """
    if df_enc.empty:
        ws.write(0, 0, "Nenhum insumo encontrado na encomenda para os códigos informados.")
        return

    header_style = _xls_header_style()
    price_style  = _xls_style("#,##0.00")
    date_style   = _xls_style("DD/MM/YYYY")
    text_style   = _xls_style()

    for col_idx, (label, _, _) in enumerate(layout):
        ws.write(0, col_idx, label, header_style)
    ws.row(0).height_mismatch = True
    ws.row(0).height = 20 * 20  # twips (1pt = 20 twips)

    max_lens = [len(label) for label, _, _ in layout]

    for row_idx, (_, row) in enumerate(df_enc.iterrows(), start=1):
        for col_idx, (_, source, kind) in enumerate(layout):
            value = row[source] if source is not None and source in df_enc.columns else None

            if _to_blank(value):
                continue
            if kind == "price":
                try:
                    num = float(str(value).replace(",", "."))
                    ws.write(row_idx, col_idx, num, price_style)
                    cell_str = f"{num:,.2f}"
                except (ValueError, TypeError):
                    cell_str = str(value)
                    ws.write(row_idx, col_idx, cell_str, text_style)
            elif kind == "date":
                s = str(value).strip()
                if "/" in s:
                    d = pd.to_datetime(s, errors="coerce", dayfirst=True)
                else:
                    d = pd.to_datetime(s, errors="coerce")
                if pd.notna(d):
                    ws.write(row_idx, col_idx, d.to_pydatetime(), date_style)
                    cell_str = d.strftime("%d/%m/%Y")
                else:
                    cell_str = ""
            else:
                cell_str = str(value)
                ws.write(row_idx, col_idx, cell_str, text_style)

            if len(cell_str) > max_lens[col_idx]:
                max_lens[col_idx] = len(cell_str)

    for col_idx, max_len in enumerate(max_lens):
        ws.col(col_idx).width = min(max_len + 4, 60) * 256

    ws.set_panes_frozen(True)
    ws.set_horz_split_pos(1)
    ws.set_remove_splits(True)


def export_carga_real(df_enc: pd.DataFrame) -> bytes:
    """Gera .xls (Excel 97-2003 / BIFF8) no layout coleta_*.xls (46 colunas,
    schema PT-BR) em uma única aba 'Carga BP', contendo todas as linhas
    recebidas (um ou vários informantes).

    Recebe um DataFrame já filtrado do encomenda_*.parquet (schema real).
    """
    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Carga BP")
    _write_carga_sheet(ws, df_enc, CARGA_REAL_COLUMNS)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()


# Layout extra para a Carga REPROVADA — mesma estrutura + colunas analíticas
# para o revisor entender por que cada linha caiu na manutenção.
CARGA_REPROVADA_EXTRA: list[tuple[str, str | None, str]] = [
    ("Motivo Manutenção",        "motivo_manutencao",   "text"),
    ("Validação Variação",       "validacao_variacao",  "text"),
    ("Validação Faltas",         "validacao_faltas",    "text"),
    ("Validação Incidência",     "validacao_incidencia", "text"),
    ("Último Preço BP",          "ultimo_preco_db",     "price"),
    ("Variação Atual (%)",       "variacao_atual_pct",  "price"),
    ("Limite Inf. Variação (%)", "vl_lim_inf_var",      "price"),
    ("Limite Sup. Variação (%)", "vl_lim_sup_var",      "price"),
    ("Faltas",                   "total_faltas_com_atual", "text"),
    ("Limite de Faltas",         "faltas_consecutivas_aceitas", "text"),
]


def export_carga_reprovada(df_enc: pd.DataFrame) -> bytes:
    """Gera .xls da carga REPROVADA (manutencao=1) — mesmo layout da carga
    real + colunas com motivo e métricas da crítica para revisão manual.
    """
    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Carga Reprovada")
    layout = CARGA_REAL_COLUMNS + CARGA_REPROVADA_EXTRA
    _write_carga_sheet(ws, df_enc, layout)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()
